import openpyxl
import string as s
import datetime
import requests
import re

read_path = 'C:/Users/akhil/Documents/PS-1/24-5-2019/puc_data.xlsx'
write_path = 'C:/Users/akhil/Documents/PS-1/24-5-2019/template.xlsx'

rd_obj = openpyxl.load_workbook(read_path)
wt_obj = openpyxl.load_workbook(write_path)

rd_sheet = rd_obj.active
wt_sheet = wt_obj.active

for i in range(1,rd_sheet.max_row+1):
    cell = rd_sheet.cell(row = i, column = 1)
    sno = cell.value
    if str(sno).isnumeric() == True:

        n = wt_sheet.max_row+1
        # Roll number to ID
        cell = rd_sheet.cell(row = i, column = 3)
        roll = cell.value
        cell = wt_sheet.cell(row = n, column = 1)
        cell.value = roll

        #  Application number to USN
        cell = rd_sheet.cell(row = i, column = 2)
        appl = cell.value
        cell = wt_sheet.cell(row = n, column = 2)
        cell.value = appl

        # Part of Section to Degree
        cell = rd_sheet.cell(row = i, column = 11)
        sectionl = cell.value
        cell = wt_sheet.cell(row = n, column = 3)
        section = sectionl[:4]
        if section == 'PCMB':
            degree = "Science"
        elif section == "HEPS":
            degree = "Arts"
        else:
            degree = "Commerce"
        cell.value = degree

        # Other part of section to Department ID
        cell = wt_sheet.cell(row = n, column = 4)
        cell.value = sectionl[:4]

        # Part of Section to Department
        cell = wt_sheet.cell(row = n, column = 5)
        cell.value = sectionl[:4]

        # Other part of sectin to section
        cell = wt_sheet.cell(row = n, column = 6)
        cell.value = sectionl

        #Name
        cell = rd_sheet.cell(row = i, column = 4)
        name = cell.value
        namel = name.split()


        sflag = False
        if '(S)' in namel:
            namel.remove('(S)')
            sflag = True
        if len(namel[0]) > 1:
            firstname = namel[0]
            lastname = ' '.join(namel[1:])
        else:

            for x in range(len(namel)):
                    if len(namel[x]) > 1:
                        firstname = namel[x]
                        lastname1 = (' ').join(namel[:x])
                        lastname2 = (' ').join(namel[x+1:])
                        lastname = lastname1 +' '+ lastname2
            

        cell = wt_sheet.cell(row = n, column = 7)
        cell.value = firstname
        cell = wt_sheet.cell(row = n, column = 8)
        cell.value = lastname

        # DOB to DOB
        cell = rd_sheet.cell(row = i, column = 24)
        dateof = str(cell.value)
        cell = wt_sheet.cell(row = n, column = 9)
        
        cell.value = (str(dateof)[:10])
        cell.number_format = 'YYYY MMM DD'

        #M/F to Gender
        cell = rd_sheet.cell(row = i, column = 9)
        genderl = str(cell.value)
        if genderl[:1].upper() == 'M':
            gender = "Male"
        else:
            gender = "Female"
        cell = wt_sheet.cell(row = n, column = 10)
        cell.value = gender

        #Phone No. to Mobile
        cell = rd_sheet.cell(row = i, column = 16)
        phone = cell.value
        cell = wt_sheet.cell(row = n, column = 11)
        cell.value = phone

        #Father's Name to Father's Name
        cell = rd_sheet.cell(row = i, column = 14)
        phone = cell.value
        cell = wt_sheet.cell(row = n, column = 13)
        cell.value = phone

        #Setting everyone's pasword as same
        cell = wt_sheet.cell(row = n, column = 29)
        password = "MyPUCollegeSucks"
        cell.value = password

        #Setting everyone's user type to student
        cell = wt_sheet.cell(row = n, column = 30)
        account = "Student"
        cell.value = account

        #Setting everyone's role to 'to study'
        cell = wt_sheet.cell(row = n, column = 31)
        role = "To study"
        cell.value = role

        #Setting ID to Role No.
        cell = wt_sheet.cell(row = n, column = 32)
        cell.value = roll

        #Batch to Academic period
        cell = rd_sheet.cell(row = i,column = 10)
        acad_period = cell.value
        cell = wt_sheet.cell(row = n, column = 34)
        cell.value = acad_period

        cell = wt_sheet.cell(row = n, column = 33)
        cell.value = acad_period
        
        #Date of joining to Year of joining
        cell = rd_sheet.cell(row = i, column = 30)
        joining_date = cell.value
        cell = wt_sheet.cell(row = n, column = 35)
        cell.value = joining_date[-4:]

        #Physically Handicap to special category
        cell = rd_sheet.cell(row = i, column = 22)
        handicap = cell.value
        cell = wt_sheet.cell(row = n, column = 37)
        cell.value = handicap

        #Caste to Basic category
        cell = rd_sheet.cell(row = i, column = 23)
        category = cell.value
        cell = wt_sheet.cell(row = n, column = 39)
        cell.value = category

        #Caste category to reserved category
        cell = rd_sheet.cell(row = i, column = 40)
        reserved = cell.value
        cell = wt_sheet.cell(row = n, column = 38)
        cell.value = reserved

        #Degree year
        cell = wt_sheet.cell(row = n, column = 41 )
        cell.value = '2017-2018'

        #Adoption
        if sflag == True:
            cell = wt_sheet.cell(row = n, column = 43 )
            cell.value = 'Sports'

        #Address fields--------------------
        cell = rd_sheet.cell(row = i, column = 15)
        address = str(cell.value)
        if len(str(address)) != 0:
            for x in range(len(address)):
                if address[x] == '\n':
                    address = address[x+1:]
                    break
        while 1:
            if address[-1] == '\n':
                address = address[0:-1]
            else:
                break
        #print(address)
        addressl = address.split('\n')
        


        # for x in addressl:
        #     match = re.search(r"\S+ TQ",x)
        #     if match:
        #         taluq = match.group(0).split()
        #         print("Taluq is:",taluq[0])
        #         break

        for x in addressl:
            match = re.search(r"\S+ DIST",x)
            if match:
                district = match.group(0).split()
                #print("District is:",district[0])
                cell = wt_sheet.cell(row = n, column = 20 )
                if district[0] != '&':
                    cell.value = str(district[0].title())
                break


        for x in addressl:
            match = re.search(r"\S+ STATE",x)
            if match:
                state = match.group(0).split()
                cell = wt_sheet.cell(row = n, column = 21 )
                if district[0] != '&':
                    cell.value = str(state[0].title())
                break

        for x in range(len(address)):
            if address[x:x+6].isnumeric() == True:
                pincode = address[x:x+6]
                print("Pincode is:",pincode)
                request = requests.get('http://postalpincode.in/api/pincode/'+str(pincode))
                if (request.json()["Status"]) != "Error":
                    for po in request.json()['PostOffice']:
                        cell = wt_sheet.cell(row = n, column = 21 )
                        cell.value = po["State"]
                        cell = wt_sheet.cell(row = n, column = 22 )
                        cell.value = po["Country"]
                        cell = wt_sheet.cell(row = n, column = 20 )
                        cell.value = po["District"]
                        cell = wt_sheet.cell(row = n, column = 18 )
                        cell.value = po["Name"]
                        break
                cell = wt_sheet.cell(row = n, column = 19 )
                cell.value = pincode
                break
    print(i)
wt_obj.save(write_path)