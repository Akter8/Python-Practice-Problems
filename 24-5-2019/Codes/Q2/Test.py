import re
import openpyxl
import string as s
import datetime
import requests

read_path = 'C:/Users/akhil/Documents/PS-1/24-5-2019/puc_data.xlsx'
write_path = 'C:/Users/akhil/Documents/PS-1/24-5-2019/template.xlsx'

rd_obj = openpyxl.load_workbook(read_path)
wt_obj = openpyxl.load_workbook(write_path)

rd_sheet = rd_obj.active
wt_sheet = wt_obj.active

count = 0
for i in range(1,rd_sheet.max_row+1):
    cell = rd_sheet.cell(row = i, column = 1)
    sno = cell.value
    if str(sno).isnumeric() == True:
        n = wt_sheet.max_row+1

        cell = rd_sheet.cell(row = i, column = 4)
        name = cell.value
        print(name)


        cell = rd_sheet.cell(row = i, column = 15)
        address = str(cell.value)
        if len(str(address)) != 0:
            for x in range(len(address)):
                if address[x] == '\n':
                    address = address[x+1:]
                    break
        while 1:
            if address[-1] == '\n':
                address = address[0:-1]
            else:
                break
        print(address)
        count += 1
        addressl = address.split('\n')
        


        for x in addressl:
            match = re.search(r"\S+ TQ",x)
            if match:
                taluq = match.group(0).split()
                print("Taluq is:",taluq[0])
                break

        for x in addressl:
            match = re.search(r"\S+ DIST",x)
            if match:
                district = match.group(0).split()
                print("District is:",district[0])
                break

        for x in range(len(address)):
            if address[x:x+6].isnumeric() == True:
                pincode = address[x:x+6]
                print("Pincode is:",pincode)
                request = requests.get('http://postalpincode.in/api/pincode/584107')
                print(request.json())
                if (request.json()["Status"]) != "Error":
                    for po in request.json()['PostOffice']:
                        print(po["Name"])
                        break
                break
        print('')

print(count)



