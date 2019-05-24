import cryptography as cpt
import string as s
import os
import colorama
import getpass
import time
import datetime as dt
import openpyxl as xl
import hashlib


def addCard(username):
    while 1:
        print("Select card type:")
        print("\t(1) Debit")
        print("\t(2) Credit")
        card_type = int(input())
        if card_type == 1 or card_type == 2:
            break
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)


    while 1:
        print("Enter card type:")
        print("\t(1) VISA")
        print("\t(2) MasterCard")
        print("\t(3) RuPay")
        print(colorama.Fore.RED+"Press 0 to start over"+colorama.Style.RESET_ALL)
        card = int(input())
        if card>=1 and card<=3:
            break
        elif card == 0:
            os.system('cls||clear')
            addCard(username)
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)


    while 1:
        print("Enter your Bank details")
        print("\t(1) State Bank of India")
        print("\t(2) ICICI Bank")
        print("\t(3) HSBC Bank")
        print("\t(4) HDFC Bank")
        print("\t(5) Kotak Mahindra")
        print("\t(6) Bank of Baroda")
        print("\t(7) Bank of Maharshtra")
        print("\t(8) Karur Vysya Bank")
        print("\t(9) Canara Bank")
        print("\t(10) Axis Bank")
        print("\t(11) Punjab National Bank")
        print("\t(12) JP Morgan Chase")
        print("\t(13) Bank of America")
        print("\t(14) Wells Fargo")
        print("\t(15) Citi Bank")
        print("\t(16) Others")
        print(colorama.Fore.RED+"Press 0 to start over"+colorama.Style.RESET_ALL)
        print("Enter the corresponding number to your bank")
        
        bank = int(input())
        if bank>=1 and bank<=16:
            break
        elif bank == 0:
            os.system('cls||clear')
            addCard(username)
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
        


    print("Enter your card number")
    card_number = (input())
    loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/user.xlsx'

    wb_obj = xl.load_workbook(loc)
    sheet = wb_obj.active

    i = 0
    present = False
    for i in range(1,sheet.max_row+1):
        number = sheet.cell(row = i,column = 5)
        if number.value == card_number:
            present = True
            break

    if present == True:
        os.system('cls||clear')
        print(colorama.Fore.RED+"This CardNumber is already present in the database"+colorama.Style.RESET_ALL)
        return

    while 1:
        print("Enter the Expiry date of your card")
        print("Year of expiry")
        year = int(input())
        print("Month of expiry")
        month = int(input())
        date = dt.date(year,month,30)
        if date < dt.date.today():
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid Date of Expiry"+colorama.Style.RESET_ALL)
            continue
        else:
            break

    os.system('cls||clear')
    print("Enter the name on the card: ")
    name = input()

    loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/info.xlsx'
    wb_obj = xl.load_workbook(loc)
    sheetr = wb_obj.active


    i = sheetr.max_row
    cell = sheetr.cell(row = i+1,column = 1)
    cell.value = username
    cell = sheetr.cell(row = i+1,column = 2)
    cell.value = card_type
    cell = sheetr.cell(row = i+1,column = 3)
    cell.value = card
    cell = sheetr.cell(row = i+1,column = 4)
    cell.value = bank
    cell = sheetr.cell(row = i+1,column = 5)
    cell.value = card_number
    cell = sheetr.cell(row = i+1,column = 6)
    cell.value = date
    cell = sheetr.cell(row = i+1,column = 7)
    cell.value = name
    

    wb_obj.save('C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/info.xlsx')




def editCards(username):
    loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/info.xlsx'
    wb_obj = xl.load_workbook(loc)
    sheet = wb_obj.active

    print("Your cards are:")
    for i in range(1, sheet.max_row+1):
        user = sheet.cell(row = i, column = 1)
        if user.value == username:
            for j in range(1,sheet.max_column+1):
                cell = sheet.cell(row = i, column = j)
                print(cell.value, end=' ')
            print(' ')

    print("\nEnter the card number of the card you want to change the details of:")
    number = int(input())

    for i in range(1, sheet.max_row+1):
        num = sheet.cell(row = i, column = 5)
        if num.value == number:
            k = i
            print("What value do you want to change?")
            print("\t(1) Card Type")
            print("\t(2) Card Company")
            print("\t(3) Bank")
            print("\t(4) Card Number")
            print("\t(5) Expiry date")
            print("\t(6) Name of user on card")
            print("Enter the corresponding number")
            n = int(input())
            if n == 1:
                os.system('cls||clear')
                while 1:
                    print("Select card type:")
                    print("\t(1) Debit")
                    print("\t(2) Credit")
                    card_type = int(input())
                    if card_type == 1 or card_type == 2:
                        break
                    else:
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
                cell = sheet.cell(row = k,column = 2)
                cell.value = card_type  
         
            elif n == 2:
                os.system('cls||clear')
                while 1:
                    print("Enter card type:")
                    print("\t(1) VISA")
                    print("\t(2) MasterCard")
                    print("\t(3) RuPay")
                    card = int(input())
                    if card>=1 and card<=3:
                        break
                    else:
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
                cell = sheet.cell(row = k,column = 3)
                cell.value = card  


            elif n == 3:
                os.system('cls||clear')
                while 1:
                    print("Enter your Bank details")
                    print("\t(1) State Bank of India")
                    print("\t(2) ICICI Bank")
                    print("\t(3) HSBC Bank")
                    print("\t(4) HDFC Bank")
                    print("\t(5) Kotak Mahindra")
                    print("\t(6) Bank of Baroda")
                    print("\t(7) Bank of Maharshtra")
                    print("\t(8) Karur Vysya Bank")
                    print("\t(9) Canara Bank")
                    print("\t(10) Axis Bank")
                    print("\t(11) Punjab National Bank")
                    print("\t(12) JP Morgan Chase")
                    print("\t(13) Bank of America")
                    print("\t(14) Wells Fargo")
                    print("\t(15) Citi Bank")
                    print("\t(16) Others")
                    print(colorama.Fore.RED+"Press 0 to start over"+colorama.Style.RESET_ALL)
                    print("Enter the corresponding number to your bank")
                    
                    bank = int(input())
                    if bank>=1 and bank<=16:
                        break
                    elif bank == 0:
                        os.system('cls||clear')
                        addCard(username)
                    else:
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
                cell = sheet.cell(row = k,column = 4)
                cell.value = bank


            elif n == 4:
                os.system('cls||clear')
                while 1:
                    print("Enter your card number")
                    card_number = (input())

                    present = False
                    for i in range(1,sheet.max_row+1):
                        number = sheet.cell(row = i,column = 5)
                        if number.value == card_number:
                            present = True
                            break

                    if present == True:
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"This CardNumber is already present in the database"+colorama.Style.RESET_ALL)
                        continue  
                    else:
                        break
                cell = sheet.cell(row = k,column = 5)
                cell.value = card_number 

            elif n == 5:               
                while 1:
                    print("Enter the Expiry date of your card")
                    print("Year of expiry")
                    year = int(input())
                    print("Month of expiry")
                    month = int(input())
                    date = dt.date(year,month,30)
                    if date < dt.date.today():
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"Please enter a valid Date of Expiry"+colorama.Style.RESET_ALL)
                        continue
                    else:
                        break
                cell = sheet.cell(row = k,column = 6)
                cell.value = date

            elif n == 6:
                print("Enter the name on the card: ")
                name = input()
                cell = sheet.cell(row = k,column = 7)
                cell.value = name

            else:
                os.system('cls||clear')
                print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
                continue
            break
        
    wb_obj.save(loc)
    return





def settings(username):
    loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/user.xlsx'
    wb_obj = xl.load_workbook(loc)
    sheet = wb_obj.active

    print("Your informations is:")
    for i in range(1,sheet.max_row+1):
        user = sheet.cell(row = i, column = 1)
        if user.value == username:
            for j in range(1,sheet.max_column+1):
                cell = sheet.cell(row = i, column = j)
                print(cell.value, end=' ')
            k = i
            break

    print("What value do you want to change?")
    print("\t(1) Password")
    print("\t(2) Email-ID")
    print("\t(3) Full Name")
    print("\t(4) Contact Number")
    print("Enter the corresponding number")
    n = int(input())
    if n == 1:
        os.system('cls||clear')
        while 1:
            # Initialisation to check the validity of the password
            small = 0
            big = 0
            number = 0
            small_alphabets = list(s.ascii_lowercase)
            big_alphabets = list(s.ascii_uppercase)
            digits = list(s.digits)

            print(colorama.Style.RESET_ALL)
            
            print("Please enter a PASSWORD of your choice")
            print("With atleast one capital letter and a number and length of at least 6 and maximum of 12 characters")
            print(colorama.Fore.RED+"Press (e/E) to exit to main screen"+colorama.Style.RESET_ALL)
            password = getpass.getpass()

            if password == 'e' or password == 'E':
                return 

            complies = False
            for letter in password:
                if letter in small_alphabets:
                    small += 1
                elif letter in big_alphabets:
                    big += 1
                elif letter in digits:
                    number += 1
                else:
                    break
            
            if small>=1 and big>=1 and number>=1 and len(password)>=6 and len(password)<=12:
                complies = True
            
            goAhead = False

            if complies == True:
                os.system('cls||clear')
                print("Please enter your PASSWORD again to CONFIRM it:")
                confirm_password = getpass.getpass()
                if confirm_password == password:
                    goAhead = True
                else:
                    os.system('cls||clear')
                    print(colorama.Fore.RED+"Your password does not match, Please re-enter your password.")
                    continue
            if goAhead == False:
                os.system('cls||clear')
                print(colorama.Fore.RED+"Entered string does not qualify as a password")
                continue
            else:
                break   

        file = open('salt.txt','rb')
        salt = file.read()
        file.close()
        password = str(hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000))


        cell = sheet.cell(row = k,column = 2)
        cell.value = password
    
    elif n == 2:
        os.system('cls||clear')
        print("Enter your EMAIL-ID")
        email = input()
        cell = sheet.cell(row = k,column = 3)
        cell.value = email
    
    elif n == 3:
        os.system('cls||clear')
        print("Enter your FULLNAME")
        name = input()
        cell = sheet.cell(row = k,column = 4)
        cell.value = name    

    elif n == 4:
        os.system('cls||clear')
        print("Enter your CONTACT NUMBER")
        phone = int(input())
        cell = sheet.cell(row = k,column = 5)
        cell.value = phone 

    else:
        os.system('cls||clear')
        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
        settings(username)
        
    wb_obj.save(loc)
    return





def showDetails(username):
    while 1:
        print(colorama.Fore.GREEN+"Welcome to your home screen",username,"!!\nWhat would you like to do today?"+colorama.Style.RESET_ALL)
        print("\t(1) Add Credit/Debit card")
        print("\t(2) Edit your existing Card information")
        print("\t(3) Change your personal information (including password)")
        print("\t(4) Logout")
        print("Please press the required button.")
        pref = int(input())

        if pref == 1:
            os.system('cls||clear')
            addCard(username)
        elif pref == 2:
            os.system('cls||clear')
            editCards(username)
        elif pref == 3:
            os.system('cls||clear')
            settings(username)
        elif pref == 4:
            return
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)