import cryptography as cpt
import string as s
import os
import colorama
import getpass
import time
import myFunctions2 as mf2
import openpyxl as xl
import hashlib


def mainScreen():
    os.system('cls||clear')
    print("*****####****####****####****####****####****####****####****####****####****####****")
    print(colorama.Fore.GREEN+"Welcome to Credit/Debit card Information CLI app"+colorama.Style.RESET_ALL)
    while 1:
        print(colorama.Style.RESET_ALL)
        print("Please decide what you want to do")
        print("Press the required button to access your preference:\n\t(1) To Create New User\n\t(2) Sign-in")
        print("\t(3) To Exit")
        pref = int(input())
        if pref == 1 or pref == 2:
            os.system('cls||clear')
            return(pref)
        elif pref == 3:
            os.system('cls||clear')
            print(colorama.Fore.GREEN+"Thank you for visiting us, please visit again\n"+colorama.Style.RESET_ALL)
            exit()
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed"+colorama.Style.RESET_ALL)



def addUser():
    print("\n\nHello there new User!!\nJust a few fields for you to fill and you'll be all set up")

    # USERNAME
    while 1:
        # Initialisations to check the validity of username
        small = 0
        big = 0
        number = 0
        small_alphabets = list(s.ascii_lowercase)
        big_alphabets = list(s.ascii_uppercase)
        digits = list(s.digits)
        
        print(colorama.Style.RESET_ALL)
        print("Please enter a USERNAME of your choice")
        print("With atleast one capital letter and a number and length of at least 6 and maximum of 12 characters")
        print(colorama.Fore.RED+"Press (e/E) to exit to main screen"+colorama.Style.RESET_ALL)

        username = input()
        
        if username == 'e' or username == 'E':
            return 

        # Checking to see if the entered string complies with the 
        # given restrictions or not
        complies = False
        random_character = False
        for letter in username:
            if letter in small_alphabets:
                small += 1
            elif letter in big_alphabets:
                big += 1
            elif letter in digits:
                number += 1
            else:
                random_character = True
                break
        
        if small>=1 and big>=1 and number>=1 and len(username)>=6 and len(username)<=12:
            complies = True
        if random_character == False and complies == True:
            pass
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid username")
            continue
                

        # If and after the username qualifies,
        # we check to see if username is already taken or not in this file
        loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/user.xlsx'
        wb_obj = xl.load_workbook(loc)
        sheet = wb_obj.active
        

        # To check if a username is already taken or not 4
        # by comparing it with every username already present
        i = 0
        taken = False
        for i in range(1, sheet.max_row+1):
            user = sheet.cell(row = i, column = 1)
            if user.value == username:
                taken = True
                break
        
        if taken == True:
            os.system('cls||clear')
            print(colorama.Fore.RED+"This username is already taken, please fill another username")
            continue
        else:
            break
    
    # PASSWORD
    os.system('cls||clear')
    while 1:
        # Initialisation to check the validity of the password
        small = 0
        big = 0
        number = 0
        small_alphabets = list(s.ascii_lowercase)
        big_alphabets = list(s.ascii_uppercase)
        digits = list(s.digits)

        print(colorama.Style.RESET_ALL)
        
        print("Please enter a PASSWORD of your choice")
        print("With atleast one capital letter and a number and length of at least 6 and maximum of 12 characters")
        print(colorama.Fore.RED+"Press (e/E) to exit to main screen"+colorama.Style.RESET_ALL)
        password = getpass.getpass()

        if password == 'e' or password == 'E':
            return 

        complies = False
        for letter in password:
            if letter in small_alphabets:
                small += 1
            elif letter in big_alphabets:
                big += 1
            elif letter in digits:
                number += 1
            else:
                random_character = True
                break
        
        if small>=1 and big>=1 and number>=1 and len(password)>=6 and len(password)<=12:
            complies = True
        
        goAhead = False

        if complies == True:
            os.system('cls||clear')
            print("Please enter your PASSWORD again to CONFIRM it:")
            confirm_password = getpass.getpass()
            if confirm_password == password:
                goAhead = True
            else:
                os.system('cls||clear')
                print(colorama.Fore.RED+"Your password does not match, Please re-enter your password.")
                continue
        if goAhead == False:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Entered string does not qualify as a password")
            continue
        else:
            break

    os.system('cls||clear')
    print("Enter your EMAIL-ID")
    print(colorama.Fore.RED+"Press (e/E) to exit to main screen"+colorama.Style.RESET_ALL)
    email = input()
    if email == 'e' or email == 'E':
        return 

    os.system('cls||clear')
    print("Enter your FULLNAME")
    print(colorama.Fore.RED+"Press (e/E) to exit to main screen"+colorama.Style.RESET_ALL)
    name = input()
    if name == 'e' or name == 'E':
        return 
    
    os.system('cls||clear')
    print("Enter your CONTACT NUMBER")
    print(colorama.Fore.RED+"Press (e/E) to exit to main screen"+colorama.Style.RESET_ALL)
    phonel = input()
    if phonel == 'e' or phonel == 'E':
        return 
    phone = int(phonel)

    
    loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/user.xlsx'
    wb_obj = xl.load_workbook(loc)
    sheet = wb_obj.active
 
    num_rows = sheet.max_row
    
    file = open('salt.txt','rb')
    salt = file.read()
    file.close()
    password = str(hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000))


    i = num_rows + 1
    cell = sheet.cell(row = i, column = 1)
    cell.value = username
    cell = sheet.cell(row = i, column = 2)
    cell.value = password
    cell = sheet.cell(row = i, column = 3)
    cell.value = email
    cell = sheet.cell(row = i, column = 4)
    cell.value = name
    cell = sheet.cell(row = i, column = 5)
    cell.value = phone

    wb_obj.save(loc)
    os.system('cls||clear')
    print(colorama.Fore.GREEN+"Done creating a new user, you can now sign in"+colorama.Style.RESET_ALL)
    time.sleep(3)
    mainScreen()


def signIn():
    while 1:
        os.system('cls||clear')
        print("Please enter your USERNAME")
        username = input()
        os.system('cls||clear')
        print("Please enter your PASSWORD")
        password = getpass.getpass()

        goAhead = False

        loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/user.xlsx'
        wb_obj = xl.load_workbook(loc)
        sheet = wb_obj.active

        
        for i in range(1,sheet.max_row+1):
            user = sheet.cell(row = i,column = 1)
            if user.value == username:
                passw = sheet.cell(row = i,column = 2)
                passw = passw.value
                file = open('salt.txt','rb')
                salt = file.read()
                file.close()
                password = str(hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000))

                if passw == password:
                    goAhead = True
                break
        
        if goAhead == False:
            os.system('cls||clear')
            print(colorama.Fore.RED+"You have entered either a wrong USERNAME or a wrong PASSWORD"+colorama.Style.RESET_ALL)
            continue
        else:
            os.system('cls||clear')
            mf2.showDetails(username)
            break