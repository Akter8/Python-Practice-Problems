import cryptography as cpt
import string as s
import os
import colorama
import getpass
import time
import datetime as dt
import openpyxl as xl
import hashlib
from pymongo import MongoClient

myclient = MongoClient("mongodb://localhost:27017/")
mydb = myclient["CreditCardCLI"]
users = mydb["users"]
info = mydb["info"]

def addCard(username):
    while 1:
        print("Select card type:")
        print("\t(1) Debit")
        print("\t(2) Credit")
        card_type = int(input())
        if card_type == 1 or card_type == 2:
            break
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)

    os.system('cls||clear')
    while 1:
        print("Enter card type:")
        print("\t(1) VISA")
        print("\t(2) MasterCard")
        print("\t(3) RuPay")
        print(colorama.Fore.RED+"Press 0 to start over"+colorama.Style.RESET_ALL)
        card = int(input())
        if card>=1 and card<=3:
            break
        elif card == 0:
            os.system('cls||clear')
            addCard(username)
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)

    os.system('cls||clear')
    while 1:
        print("Enter your Bank details")
        print("\t(1) State Bank of India")
        print("\t(2) ICICI Bank")
        print("\t(3) HSBC Bank")
        print("\t(4) HDFC Bank")
        print("\t(5) Kotak Mahindra")
        print("\t(6) Bank of Baroda")
        print("\t(7) Bank of Maharshtra")
        print("\t(8) Karur Vysya Bank")
        print("\t(9) Canara Bank")
        print("\t(10) Axis Bank")
        print("\t(11) Punjab National Bank")
        print("\t(12) JP Morgan Chase")
        print("\t(13) Bank of America")
        print("\t(14) Wells Fargo")
        print("\t(15) Citi Bank")
        print("\t(16) Others")
        print(colorama.Fore.RED+"Press 0 to start over"+colorama.Style.RESET_ALL)
        print("Enter the corresponding number to your bank")
        
        bank = int(input())
        if bank>=1 and bank<=16:
            break
        elif bank == 0:
            os.system('cls||clear')
            addCard(username)
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
        

    os.system('cls||clear')
    print("Enter your card number")
    card_number = (input())
    #loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/user.xlsx'

    #wb_obj = xl.load_workbook(loc)
    #sheet = wb_obj.active

    i = 0
    present = False

    for infor in info.find():
        if infor["CardNumber"] == card_number:
            present = True
            break

    # for i in range(1,sheet.max_row+1):
    #     number = sheet.cell(row = i,column = 5)
    #     if number.value == card_number:
    #         present = True
    #         break

    if present == True:
        os.system('cls||clear')
        print(colorama.Fore.RED+"This CardNumber is already present in the database"+colorama.Style.RESET_ALL)
        return


    print("Enter the Expiry date of your card")
    print("Year of expiry")
    year = int(input())

    os.system('cls||clear')
    print("Enter the name on the card: ")
    name = input()

    # loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/info.xlsx'
    # wb_obj = xl.load_workbook(loc)
    # sheetr = wb_obj.active

    
    new_row = {"Username":username, "CardType":card_type, "Card":card, "Bank":bank, "CardNumber":card_number, "ExpiryDate":year, "NameOnCard":name}
    info.insert_one(new_row)
    #wb_obj.save('C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/info.xlsx')




def editCards(username):
    # loc = 'C:/Users/akhil/Documents/PS-1/23-5-2019/CLI/info.xlsx'
    # wb_obj = xl.load_workbook(loc)
    # sheet = wb_obj.active

    print("Your cards are:")
    count = 1
    for infor in info.find():
        
        if infor["Username"] == username:
            print("Card:",count)
            count += 1
            #print("\tUsername:",infor["Username"],"\n\tCard type:",infor["CardType"],"\n\tCard info:",infor["Card"],"\n\tBank:",infor["Bank"],"\n\tCard Number:",infor["CardNumber"],"\n\tExpiry Date:",infor["ExpiryDate"],"\n\tName on the card:",infor["NameOnCard"])
            print("\tUsername:",infor["Username"])
            if int(infor["CardType"]) == 1:
                cardtype = "Debit"
            elif int(infor["CardType"]) == 2:
                cardtype = "Credit"
            print("\tCard Type:",cardtype)

            if int(infor["Card"]) == 1:
                cardtype = "VISA"
            elif int(infor["Card"]) == 2:
                cardtype = "MasterCard"
            elif int(infor["CardType"]) == 3:
                cardtype = "RuPay"
            print("\tCard Info:",cardtype)

            print("\tCard Number:",infor["CardNumber"])
            
            if int(infor["Bank"]) == 1:
                bank = "State Bank of India"
            elif int(infor["Bank"]) == 2:
                bank = "ICICI Bank"
            elif int(infor["Bank"]) == 3:
                bank = "HSBC Bank"
            elif int(infor["Bank"]) == 4:
                bank = "HDFC Bank"
            elif int(infor["Bank"]) == 5:
                bank = "Kotak Mahindra"
            elif int(infor["Bank"]) == 6:
                bank = "Bank of Baroda"
            elif int(infor["Bank"]) == 7:
                bank = "Bank of Maharashtra"
            elif int(infor["Bank"]) == 8:
                bank = "Karur Vysya Bank"
            elif int(infor["Bank"]) == 9:
                bank = "Canara Bank"
            elif int(infor["Bank"]) == 10:
                bank = "Axis Bank"
            elif int(infor["Bank"]) == 11:
                bank = "Punjab National Bank"
            elif int(infor["Bank"]) == 12:
                bank = "JP Morgan Chase"
            elif int(infor["Bank"]) == 13:
                bank = "Bank of America"
            elif int(infor["Bank"]) == 14:
                bank = "Wells Fargo"
            elif int(infor["Bank"]) == 15:
                bank = "Citi Bank"
            elif int(infor["Bank"]) == 16:
                bank = "Not Specified"
            print("\tBank:",bank)

            print("\tYear of Expiry:",infor["ExpiryDate"])
            print("\tName on the Card",infor["NameOnCard"])
            

    print("\nEnter the card number of the card you want to change the details of:")
    number = (input())
    noMatch = True
    for infor in info.find():
        #print(infor["CardNumber"])
        if int(number) == int(infor["CardNumber"]) and username == infor["Username"]:
            noMatch = False
            print("What value do you want to change?")
            print("\t(1) Card Type")
            print("\t(2) Card Company")
            print("\t(3) Bank")
            print("\t(4) Card Number")
            print("\t(5) Expiry date")
            print("\t(6) Name of user on card")
            print("\t(7) Delete the card information")
            print("Enter the corresponding number")
            n = int(input())
            if n == 1:
                os.system('cls||clear')
                while 1:
                    print("Select card type:")
                    print("\t(1) Debit")
                    print("\t(2) Credit")
                    card_type = int(input())
                    if card_type == 1 or card_type == 2:
                        break
                    else:
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
                # cell = sheet.cell(row = k,column = 2)
                # cell.value = card_type
                query = {"CardNumber":number} 
                new_query = {"$set":{"CardType":card_type}} 
                info.update_one(query,new_query)
         
            elif n == 2:
                os.system('cls||clear')
                while 1:
                    print("Enter card type:")
                    print("\t(1) VISA")
                    print("\t(2) MasterCard")
                    print("\t(3) RuPay")
                    card = int(input())
                    if card>=1 and card<=3:
                        break
                    else:
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
                # cell = sheet.cell(row = k,column = 3)
                # cell.value = card 
                query = {"CardNumber":number} 
                new_query = {"$set":{"Card":card}} 
                info.update_one(query,new_query) 


            elif n == 3:
                os.system('cls||clear')
                while 1:
                    print("Enter your Bank details")
                    print("\t(1) State Bank of India")
                    print("\t(2) ICICI Bank")
                    print("\t(3) HSBC Bank")
                    print("\t(4) HDFC Bank")
                    print("\t(5) Kotak Mahindra")
                    print("\t(6) Bank of Baroda")
                    print("\t(7) Bank of Maharshtra")
                    print("\t(8) Karur Vysya Bank")
                    print("\t(9) Canara Bank")
                    print("\t(10) Axis Bank")
                    print("\t(11) Punjab National Bank")
                    print("\t(12) JP Morgan Chase")
                    print("\t(13) Bank of America")
                    print("\t(14) Wells Fargo")
                    print("\t(15) Citi Bank")
                    print("\t(16) Others")
                    print(colorama.Fore.RED+"Press 0 to start over"+colorama.Style.RESET_ALL)
                    print("Enter the corresponding number to your bank")
                    
                    bank = int(input())
                    if bank>=1 and bank<=16:
                        break
                    elif bank == 0:
                        os.system('cls||clear')
                        addCard(username)
                    else:
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
                # cell = sheet.cell(row = k,column = 4)
                # cell.value = bank
                query = {"CardNumber":number} 
                new_query = {"$set":{"Bank":bank}} 
                info.update_one(query,new_query)


            elif n == 4:
                os.system('cls||clear')
                while 1:
                    print("Enter your card number")
                    card_number = (input())

                    present = False
                    for i in info.find():
                        if i["CardNumber"] == card_number:
                            present = True
                            break

                    # for i in range(1,sheet.max_row+1):
                    #     number = sheet.cell(row = i,column = 5)
                    #     if number.value == card_number:
                    #         present = True
                    #         break

                    if present == True:
                        os.system('cls||clear')
                        print(colorama.Fore.RED+"This CardNumber is already present in the database"+colorama.Style.RESET_ALL)
                        continue  
                    else:
                        break
                # cell = sheet.cell(row = k,column = 5)
                # cell.value = card_number 
                query = {"CardNumber":number} 
                new_query = {"$set":{"CardNumber":card_number}} 
                info.update_one(query,new_query)

            elif n == 5:               
                
                print("Enter the Expiry date of your card")
                print("Year of expiry")
                year = int(input())
                    # print("Month of expiry")
                    # month = int(input())
                    # date = dt.date(year,month,30)
                    # if date < dt.date.today():
                    #     os.system('cls||clear')
                    #     print(colorama.Fore.RED+"Please enter a valid Date of Expiry"+colorama.Style.RESET_ALL)
                    #     continue
                    # else:
                    #     break

                query = {"CardNumber":number} 
                new_query = {"$set":{"ExpiryDate":year}} 
                info.update_one(query,new_query)


            elif n == 6:
                print("Enter the name on the card: ")
                name = input()
                # cell = sheet.cell(row = k,column = 7)
                # cell.value = name
                query = {"CardNumber":number} 
                new_query = {"$set":{"NameOnCard":name}} 
                info.update_one(query,new_query)
            
            elif n == 7:
                os.system('cls||clear')
                print(colorama.Fore.RED+"Are you sure you want to remove this card?\nEnter the card number to confirm:"+colorama.Style.RESET_ALL)
                num = input()
                if str(number) == str(num):
                    print(colorama.Fore.GREEN+"Card Deleted"+colorama.Style.RESET_ALL)
                    info.remove({"CardNumber":number},True)
                    time.sleep(3)
                    os.system('cls||clear')
                else:
                    os.system('cls||clear')
                    print(colorama.Fore.RED+"Wrong number entered, card NOT deleted"+colorama.Style.RESET_ALL)


            else:
                os.system('cls||clear')
                print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
                continue
            break

    if noMatch == True:
        os.system('cls||clear')
        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
        editCards(username)

    return





def settings(username):
    

    print("Your informations is:")

    for user in users.find():
        if user["User"] == username:
            print("Username:",user["User"],"\nEmail:",user["Email"],"\nName:",user["Name"],"\nPhone:",user["PhoneNumber"])
            break
    

    print("What value do you want to change?")
    print("\t(1) Password")
    print("\t(2) Email-ID")
    print("\t(3) Full Name")
    print("\t(4) Contact Number")
    print("Enter the corresponding number")
    n = int(input())
    if n == 1:
        os.system('cls||clear')
        while 1:
            # Initialisation to check the validity of the password
            small = 0
            big = 0
            number = 0
            small_alphabets = list(s.ascii_lowercase)
            big_alphabets = list(s.ascii_uppercase)
            digits = list(s.digits)

            print(colorama.Style.RESET_ALL)
            
            print("Please enter a PASSWORD of your choice")
            print("With atleast one capital letter and a number and length of at least 6 and maximum of 12 characters")
            print(colorama.Fore.RED+"Press (e/E) to exit to main screen"+colorama.Style.RESET_ALL)
            password = getpass.getpass()

            if password == 'e' or password == 'E':
                return 

            complies = False
            for letter in password:
                if letter in small_alphabets:
                    small += 1
                elif letter in big_alphabets:
                    big += 1
                elif letter in digits:
                    number += 1
                else:
                    break
            
            if small>=1 and big>=1 and number>=1 and len(password)>=6 and len(password)<=12:
                complies = True
            
            goAhead = False

            if complies == True:
                os.system('cls||clear')
                print("Please enter your PASSWORD again to CONFIRM it:")
                confirm_password = getpass.getpass()
                if confirm_password == password:
                    goAhead = True
                else:
                    os.system('cls||clear')
                    print(colorama.Fore.RED+"Your password does not match, Please re-enter your password.")
                    continue
            if goAhead == False:
                os.system('cls||clear')
                print(colorama.Fore.RED+"Entered string does not qualify as a password")
                continue
            else:
                break   

        # file = open('salt.txt','rb')
        # salt = file.read()
        # file.close()
        # password = str(hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000))


        # cell = sheet.cell(row = k,column = 2)
        # cell.value = password
        query = {"User":username}
        new_query ={"$set":{"Password":password}}
        users.update_one(query,new_query)
    
    elif n == 2:
        os.system('cls||clear')
        print("Enter your EMAIL-ID")
        email = input()
        # cell = sheet.cell(row = k,column = 3)
        # cell.value = email
        query = {"User":username}
        new_query ={"$set":{"Email":email}}
        users.update_one(query,new_query)
    
    elif n == 3:
        os.system('cls||clear')
        print("Enter your FULLNAME")
        name = input()
        # cell = sheet.cell(row = k,column = 4)
        # cell.value = name  
        query = {"User":username}
        new_query ={"$set":{"Name":name}}
        users.update_one(query,new_query)  

    elif n == 4:
        os.system('cls||clear')
        print("Enter your CONTACT NUMBER")
        phone = int(input())
        # cell = sheet.cell(row = k,column = 5)
        # cell.value = phone
        query = {"User":username}
        new_query ={"$set":{"PhoneNumber":phone}}
        users.update_one(query,new_query) 

    else:
        os.system('cls||clear')
        print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)
        settings(username)
        
    #wb_obj.save(loc)
    return





def showDetails(username):
    while 1:
        print(colorama.Fore.GREEN+"Welcome to your home screen",username,"!!\nWhat would you like to do today?"+colorama.Style.RESET_ALL)
        print("\t(1) Add Credit/Debit card")
        print("\t(2) Edit your existing Card information")
        print("\t(3) Change your personal information (including password)")
        print("\t(4) View your Card information")
        print("\t(5) View your personal information")
        print("\t(6) Logout")
        print("Please press the required button.")
        pref = int(input())

        if pref == 1:
            os.system('cls||clear')
            addCard(username)
        elif pref == 2:
            os.system('cls||clear')
            editCards(username)
        elif pref == 3:
            os.system('cls||clear')
            settings(username)


        elif pref == 4:
            os.system('cls||clear')
            print("Your cards are:")
            count = 1
            for infor in info.find():
                
                if infor["Username"] == username:
                    print("Card:",count)
                    count += 1
                    #print("\tUsername:",infor["Username"],"\n\tCard type:",infor["CardType"],"\n\tCard info:",infor["Card"],"\n\tBank:",infor["Bank"],"\n\tCard Number:",infor["CardNumber"],"\n\tExpiry Date:",infor["ExpiryDate"],"\n\tName on the card:",infor["NameOnCard"])
                    print("\tUsername:",infor["Username"])
                    if int(infor["CardType"]) == 1:
                        cardtype = "Debit"
                    elif int(infor["CardType"]) == 2:
                        cardtype = "Credit"
                    print("\tCard Type:",cardtype)

                    if int(infor["Card"]) == 1:
                        cardtype = "VISA"
                    elif int(infor["Card"]) == 2:
                        cardtype = "MasterCard"
                    elif int(infor["CardType"]) == 3:
                        cardtype = "RuPay"
                    print("\tCard Info:",cardtype)

                    print("\tCard Number:",infor["CardNumber"])
                    
                    if int(infor["Bank"]) == 1:
                        bank = "State Bank of India"
                    elif int(infor["Bank"]) == 2:
                        bank = "ICICI Bank"
                    elif int(infor["Bank"]) == 3:
                        bank = "HSBC Bank"
                    elif int(infor["Bank"]) == 4:
                        bank = "HDFC Bank"
                    elif int(infor["Bank"]) == 5:
                        bank = "Kotak Mahindra"
                    elif int(infor["Bank"]) == 6:
                        bank = "Bank of Baroda"
                    elif int(infor["Bank"]) == 7:
                        bank = "Bank of Maharashtra"
                    elif int(infor["Bank"]) == 8:
                        bank = "Karur Vysya Bank"
                    elif int(infor["Bank"]) == 9:
                        bank = "Canara Bank"
                    elif int(infor["Bank"]) == 10:
                        bank = "Axis Bank"
                    elif int(infor["Bank"]) == 11:
                        bank = "Punjab National Bank"
                    elif int(infor["Bank"]) == 12:
                        bank = "JP Morgan Chase"
                    elif int(infor["Bank"]) == 13:
                        bank = "Bank of America"
                    elif int(infor["Bank"]) == 14:
                        bank = "Wells Fargo"
                    elif int(infor["Bank"]) == 15:
                        bank = "Citi Bank"
                    elif int(infor["Bank"]) == 16:
                        bank = "Not Specified"
                    print("\tBank:",bank)

                    print("\tYear of Expiry:",infor["ExpiryDate"])
                    print("\tName on the Card",infor["NameOnCard"])
                    
            print(colorama.Fore.RED+"Please presss enter to go back\n"+colorama.Style.RESET_ALL)
            lololol = input()
            os.system('cls||clear')


        elif pref == 5:
            os.system('cls||clear')
            for user in users.find():
                if user["User"] == username:
                    print("Username:",user["User"],"\nEmail:",user["Email"],"\nName:",user["Name"],"\nPhone:",user["PhoneNumber"])
                    break
            print(colorama.Fore.RED+"Please presss enter to go back\n"+colorama.Style.RESET_ALL)
            lolol = input()
            os.system('cls||clear')


        elif pref == 6:
            return
        else:
            os.system('cls||clear')
            print(colorama.Fore.RED+"Please enter a valid input to proceed\n"+colorama.Style.RESET_ALL)